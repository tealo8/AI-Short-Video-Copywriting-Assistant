# -*- coding: utf-8 -*-
"""批量导入文件解析：TXT / CSV / Excel(xlsx) 统一提取主题清单。"""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from app.core.exceptions import ParamError
from app.config import settings


def _clean_lines(lines: list[str]) -> list[str]:
    seen, result = set(), []
    for ln in lines:
        ln = ln.strip().strip("\ufeff").rstrip("。；;")
        if not ln or ln.startswith("#"):
            continue
        if ln in seen:
            continue
        seen.add(ln)
        result.append(ln)
    if len(result) > settings.BATCH_ITEM_LIMIT:
        raise ParamError(f"文件包含 {len(result)} 条主题，超过单次上限 {settings.BATCH_ITEM_LIMIT} 条，请拆分上传")
    return result


def parse_batch_file(filename: str, content: bytes) -> list[str]:
    name = (filename or "").lower()
    raw = content

    if name.endswith(".txt"):
        text = None
        for enc in ("utf-8", "gbk", "gb18030"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ParamError("TXT 文件编码无法识别，请使用 UTF-8 编码")
        lines = text.splitlines()
    elif name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        lines = []
        for row in reader:
            if row and row[0].strip():
                lines.append(row[0])
    elif name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        lines = []
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None and str(row[0]).strip():
                lines.append(str(row[0]))
        wb.close()
    else:
        raise ParamError("仅支持 TXT / CSV / Excel(.xlsx) 文件")

    result = _clean_lines(lines)
    if not result:
        raise ParamError("未解析到有效主题，请检查文件格式（每行一个主题）")
    return result
